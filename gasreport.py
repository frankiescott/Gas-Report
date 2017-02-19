import openpyxl
import smtplib
import sys
import os.path
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

#load the excel file
filepath = sys.argv[1]
wb = openpyxl.load_workbook(filepath)
ws = wb.active

#iterate through the top cell of each column to extract the names
data = []
counter = 0
for col in ws.iter_cols(max_row = 1):
    for cell in col:
        if cell.value is None:
            break
        else:
            data.append([])
            data[counter].append(cell.value)
            counter += 1

#iterate through the remaining cells in each column to extract the data
month_total, month_receipts, trucks_used, counter = 0, 0, 0, 0
for col in ws.iter_cols(min_row = 2):
    total = 0
    receipts = 0
    for cell in col:
        if cell.value is None:
            break
        else:
            total += cell.value
            receipts += 1

    if total > 0:
        data[counter].extend([total, receipts, total / receipts])
        month_total += total       #accumulate expenses
        month_receipts += receipts #accumulate receipts
        trucks_used += 1
    else:
        data[counter] += 3 * [0]

    counter += 1

#quick sort the data putting the highest expense first
def sort(data, lowIndex, highIndex):
    i = lowIndex
    j = highIndex
    pivot = data[(i + j) // 2][1]

    while i <= j:
        while data[i][1] > pivot:
            i += 1

        while data[j][1] < pivot:
            j -= 1

        if i <= j:
            data[i], data[j] = data[j], data[i]
            i += 1
            j -= 1

    if lowIndex < i - 1:
        sort(data, lowIndex, i - 1)
    if highIndex > i:
        sort(data, i, highIndex)

sort(data, 0, len(data) - 1)


#display data
basename = os.path.basename(filepath)    #
os.path.splitext(basename)               # gets the filename without the extension
filename = os.path.splitext(basename)[0] #

report_file = filename + " gas report.txt"
file = open(report_file, 'w')

file.write("{}{}{}".format("Homestead Roofing gas expenses: ", filename, "\n\n"))
file.write("{:>3}{:>20}{:>20}{:>20}{:>20}{}".format("#", "Name", "Total", "Receipts", "Average", "\n"))

for x in range(0, trucks_used):
    file.write("-"*85 + "\n")
    file.write("{:>3}{:>20}{:>20.2f}{:>20}{:>20.2f}{}".format(x + 1, data[x][0], data[x][1], data[x][2], data[x][3], "\n"))

file.write("{}{}{}{}{}{}".format("\n", month_receipts, " receipts among ", trucks_used, " trucks were logged this month.", "\n"))
file.write("{}{:.2f}{}".format("Average expense per truck: $", month_total / trucks_used, "\n"))
file.write("{}{}{}".format("Grand total: $", month_total, "\n\n"))

file.write("Trucks with no logged receipts:\n\n")
file.write("{:>3}{:>20}{:>20}{:>20}{:>20}{}".format("#", "Name", "Total", "Receipts", "Average", "\n"))
for x in range(trucks_used, len(data)):
    file.write("-"*85 + "\n")
    file.write("{:>3}{:>20}{:>20}{:>20}{:>20}{}".format(x + 1, data[x][0], "-", "-", "-", "\n"))

file.close()

#send e-mail attachment
fromaddr = ""
passwd = ""
toaddr = fromaddr
msg = MIMEMultipart()
msg['Subject'] = "Homestead Roofing gas report: " + filename

body = "Attached is the gas report for " + filename + "\n\n-Frankie"
msg.attach(MIMEText(body, 'plain'))

attachment = open(report_file, "rb")
part = MIMEBase('application', 'octet-stream')
part.set_payload((attachment).read())
encoders.encode_base64(part)
part.add_header('Content-Disposition', "attachment; filename= %s" % report_file)

msg.attach(part)

server = smtplib.SMTP('smtp.gmail.com', 587)
server.starttls()
server.login(fromaddr, passwd)

text = msg.as_string()

server.sendmail(fromaddr, fromaddr, text)
server.quit()
